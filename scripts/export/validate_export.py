"""Validate QA-AI XLSX/CSV exports against their canonical Markdown source and sidecar metadata."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path

EXPORT_DIR = Path(__file__).resolve().parent
if str(EXPORT_DIR) not in sys.path:
    sys.path.insert(0, str(EXPORT_DIR))
from parse_testcases import parse as parse_testcases
from parse_coverage import parse as parse_coverage
from parse_regression import parse as parse_regression

PARSERS = {"test-cases": parse_testcases, "coverage-review": parse_coverage, "regression-analysis": parse_regression}

TC_HEADERS = ["Test Case ID", "Module / Function", "Scenario ID", "Test Case Title", "Preconditions / Setup", "Test Steps", "Test Data", "Expected Result", "Priority", "Traceability"]
TC_KEYS = ["test_case_id", "module_function", "scenario_id", "title", "preconditions_setup", "steps", "test_data", "expected_result", "priority", "traceability"]

COVERAGE_HEADERS = ["Coverage Finding ID", "Coverage Status", "Related Source", "Current Evidence", "Finding", "Priority", "Recommended Owning Action"]
COVERAGE_KEYS = ["coverage_finding_id", "coverage_status", "related_source", "current_evidence", "finding", "priority", "recommended_owning_action"]

REGRESSION_HEADERS = ["Impact ID", "Area / Module", "Change Relationship", "Regression Scope / Behavior to Revalidate", "Impact Type", "Evidence / Traceability", "Priority", "Existing Coverage Reference", "Decision"]
REGRESSION_KEYS = ["impact_id", "area_module", "change_relationship", "behavior_to_revalidate", "impact_type", "evidence_traceability", "priority", "existing_coverage_reference", "decision"]

TIER_TITLES = {
    "Minimum / Release-Gate Regression": "minimum_release_gate",
    "Recommended Regression": "recommended",
    "Full Changed-Feature Verification": "full_changed_feature",
}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _expected(model: dict[str, object]) -> tuple[list[str], list[list[str]]]:
    artifact_type = model["artifact_type"]
    if artifact_type == "test-cases":
        return TC_HEADERS, [[_norm("\n".join(r[k]) if k == "steps" else r[k]) for k in TC_KEYS] for r in model["records"]]
    if artifact_type == "coverage-review":
        return COVERAGE_HEADERS, [[_norm(r.get(k)) for k in COVERAGE_KEYS] for r in model["records"]]
    return REGRESSION_HEADERS, [[_norm(r.get(k)) for k in REGRESSION_KEYS] for r in model["impact_records"]]


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    return (rows[0], rows[1:]) if rows else ([], [])


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]], dict[str, list[str]] | None]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX validation") from exc
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [_norm(v) for v in rows[0]] if rows else []
    data_rows = [[_norm(v) for v in row] for row in rows[1:]] if rows else []
    tiers: dict[str, list[str]] | None = None
    if "Regression Scope" in wb.sheetnames:
        tiers = {value: [] for value in TIER_TITLES.values()}
        scope = wb["Regression Scope"]
        scope_rows = list(scope.iter_rows(values_only=True))
        if not scope_rows or [_norm(v) for v in scope_rows[0][:2]] != ["Tier", "Reference ID"]:
            raise ValueError("Regression Scope sheet header mismatch; expected ['Tier', 'Reference ID']")
        for row in scope_rows[1:]:
            title = _norm(row[0] if len(row) > 0 else None)
            ref = _norm(row[1] if len(row) > 1 else None)
            if title in TIER_TITLES and ref:
                tiers[TIER_TITLES[title]].append(ref)
    return headers, data_rows, tiers


def validate(source: Path, export_path: Path, artifact_type: str) -> list[str]:
    errors: list[str] = []
    model = PARSERS[artifact_type](source)
    sidecar = export_path.with_suffix(export_path.suffix + ".export.json")
    if not sidecar.is_file():
        return [f"missing export metadata sidecar: {sidecar}"]
    metadata = json.loads(sidecar.read_text(encoding="utf-8"))
    if metadata.get("source_checksum") != sha256(source):
        errors.append("export is stale: source checksum no longer matches export metadata")
    if metadata.get("artifact_type") != artifact_type:
        errors.append("sidecar artifact_type mismatch")
    expected_headers, expected_rows = _expected(model)
    actual_tiers = None
    if export_path.suffix.lower() == ".csv":
        actual_headers, actual_rows = _read_csv(export_path)
    elif export_path.suffix.lower() == ".xlsx":
        actual_headers, actual_rows, actual_tiers = _read_xlsx(export_path)
    else:
        return ["unsupported export format; expected .csv or .xlsx"]
    if actual_headers != expected_headers:
        errors.append(f"header mismatch: expected={expected_headers} actual={actual_headers}")
    normalized_actual = [[_norm(v) for v in row] for row in actual_rows]
    if len(expected_rows) != len(normalized_actual):
        errors.append(f"record count mismatch: source={len(expected_rows)} export={len(normalized_actual)}")
    if expected_rows != normalized_actual:
        for idx, (exp, act) in enumerate(zip(expected_rows, normalized_actual), start=1):
            if exp != act:
                errors.append(f"semantic row mismatch at exported row {idx + 1}")
                break
    if artifact_type == "test-cases" and expected_headers and actual_headers:
        source_ids = [row[0] for row in expected_rows]
        export_ids = [row[0] for row in normalized_actual]
        if len(set(export_ids)) != len(export_ids):
            errors.append("duplicate Test Case ID found in export")
        if set(source_ids) != set(export_ids):
            errors.append("Test Case ID set mismatch between source and export")
    if artifact_type == "coverage-review" and expected_headers and actual_headers:
        source_ids = [row[0] for row in expected_rows]
        export_ids = [row[0] for row in normalized_actual]
        if len(set(export_ids)) != len(export_ids):
            errors.append("duplicate Coverage Finding ID found in export")
        if set(source_ids) != set(export_ids):
            errors.append("Coverage Finding ID set mismatch between source and export")
    if artifact_type == "regression-analysis" and expected_headers and actual_headers:
        source_ids = [row[0] for row in expected_rows]
        export_ids = [row[0] for row in normalized_actual]
        if len(set(export_ids)) != len(export_ids):
            errors.append("duplicate Impact ID found in export")
        if set(source_ids) != set(export_ids):
            errors.append("Impact ID set mismatch between source and export")
    if artifact_type == "regression-analysis" and export_path.suffix.lower() == ".xlsx":
        if actual_tiers is None:
            errors.append("Regression Scope sheet is missing")
        else:
            for key, expected in model["tiers"].items():
                actual = actual_tiers.get(key, [])
                if actual != expected:
                    errors.append(f"regression tier mismatch for {key}: source={expected} export={actual}")
    if metadata.get("record_count") != len(expected_rows):
        errors.append("sidecar record_count does not reconcile with canonical source")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source")
    parser.add_argument("export")
    parser.add_argument("--type", required=True, choices=tuple(PARSERS))
    args = parser.parse_args()
    if args.type == "regression-analysis" and Path(args.export).suffix.lower() == ".csv":
        print("ERROR regression-analysis round-trip validation requires XLSX because canonical tier membership is multi-table data")
        return 1
    try:
        errors = validate(Path(args.source), Path(args.export), args.type)
    except Exception as exc:
        print(f"ERROR {exc}")
        return 1
    if errors:
        for error in errors:
            print(f"ERROR {error}")
        print(f"Export validation failed; issues={len(errors)}")
        return 1
    print(f"PASS export validation: {args.export}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
