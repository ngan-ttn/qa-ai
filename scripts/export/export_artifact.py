"""Export canonical QA-AI Markdown artifacts to generic XLSX or CSV with provenance metadata."""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

EXPORT_DIR = Path(__file__).resolve().parent
ROOT = EXPORT_DIR.parents[1]
if str(EXPORT_DIR) not in sys.path:
    sys.path.insert(0, str(EXPORT_DIR))
from parse_testcases import parse as parse_testcases
from parse_coverage import parse as parse_coverage
from parse_regression import parse as parse_regression

PARSERS = {"test-cases": parse_testcases, "coverage-review": parse_coverage, "regression-analysis": parse_regression}

TC_HEADERS = ["Test Case ID", "Module / Function", "Scenario ID", "Test Case Title", "Preconditions / Setup", "Test Steps", "Test Data", "Expected Result", "Priority", "Traceability"]
TC_KEYS = ["test_case_id", "module_function", "scenario_id", "title", "preconditions_setup", "steps", "test_data", "expected_result", "priority", "traceability"]

COVERAGE_HEADERS = ["Coverage Finding ID", "Coverage Status", "Related Source", "Current Evidence", "Finding", "Priority", "Recommended Owning Action"]
COVERAGE_KEYS = ["coverage_finding_id", "coverage_status", "related_source", "current_evidence", "finding", "priority", "recommended_owning_action"]

REGRESSION_HEADERS = ["Impact ID", "Area / Module", "Change Relationship", "Regression Scope / Behavior to Revalidate", "Impact Type", "Evidence / Traceability", "Priority", "Existing Coverage Reference", "Decision"]
REGRESSION_KEYS = ["impact_id", "area_module", "change_relationship", "behavior_to_revalidate", "impact_type", "evidence_traceability", "priority", "existing_coverage_reference", "decision"]

COLUMN_WIDTHS = {
    "Test Case ID": 16, "Module / Function": 24, "Scenario ID": 18, "Test Case Title": 36,
    "Preconditions / Setup": 40, "Test Steps": 55, "Test Data": 36, "Expected Result": 55,
    "Priority": 14, "Traceability": 34,
    "Coverage Finding ID": 20, "Coverage Status": 20, "Related Source": 38, "Current Evidence": 38,
    "Finding": 55, "Recommended Owning Action": 55,
    "Impact ID": 14, "Area / Module": 28, "Change Relationship": 24,
    "Regression Scope / Behavior to Revalidate": 60, "Impact Type": 18, "Evidence / Traceability": 40,
    "Existing Coverage Reference": 42, "Decision": 22,
    "Tier": 38, "Reference ID": 22,
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def git_revision() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return "unknown"


def _string(value: object) -> str:
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return "" if value is None else str(value)


def _record_rows(records: list[dict[str, object]], keys: list[str]) -> list[list[str]]:
    return [[_string(record.get(key)) for key in keys] for record in records]


def _rows(model: dict[str, object]) -> tuple[list[str], list[list[str]], str]:
    kind = model["artifact_type"]
    if kind == "test-cases":
        return TC_HEADERS, _record_rows(model["records"], TC_KEYS), "Test Cases"
    if kind == "coverage-review":
        return COVERAGE_HEADERS, _record_rows(model["records"], COVERAGE_KEYS), "Coverage Review"
    return REGRESSION_HEADERS, _record_rows(model["impact_records"], REGRESSION_KEYS), "Regression Impact"


def _write_csv(model: dict[str, object], output: Path) -> None:
    headers, rows, _ = _rows(model)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)


def _format_sheet(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=index).column_letter
        if header in COLUMN_WIDTHS:
            ws.column_dimensions[letter].width = COLUMN_WIDTHS[header]
            continue
        values = [ws.cell(row=row, column=index).value for row in range(1, ws.max_row + 1)]
        width = max((len(str(value or "").split("\n")[0]) for value in values), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 50)


def _write_xlsx(model: dict[str, object], output: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    headers, rows, sheet = _rows(model)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _format_sheet(ws, headers)

    if model["artifact_type"] == "regression-analysis":
        tier_headers = ["Tier", "Reference ID"]
        tier_ws = wb.create_sheet("Regression Scope")
        tier_ws.append(tier_headers)
        titles = {
            "minimum_release_gate": "Minimum / Release-Gate Regression",
            "recommended": "Recommended Regression",
            "full_changed_feature": "Full Changed-Feature Verification",
        }
        for key, title in titles.items():
            for reference in model["tiers"][key]:
                tier_ws.append([title, reference])
        _format_sheet(tier_ws, tier_headers)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def export(source: Path, artifact_type: str, fmt: str, output: Path, profile: str) -> Path:
    if artifact_type == "regression-analysis" and fmt == "csv":
        raise ValueError("regression-analysis requires XLSX in the Phase 17 baseline because tier membership is multi-table data")
    model = PARSERS[artifact_type](source)
    _write_csv(model, output) if fmt == "csv" else _write_xlsx(model, output)
    count = len(model.get("records", model.get("impact_records", [])))
    metadata = {
        "schema_version": "1.0", "artifact_type": artifact_type,
        "source_path": source.as_posix(), "source_checksum": model["source_checksum"],
        "export_path": output.as_posix(), "export_format": fmt, "export_profile": profile,
        "record_count": count, "exported_at": now_iso(), "framework_revision": git_revision(),
    }
    sidecar = output.with_suffix(output.suffix + ".export.json")
    sidecar.write_text(json.dumps(metadata, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Exported {count} {artifact_type} record(s) -> {output}")
    print(f"Metadata -> {sidecar}")
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input")
    parser.add_argument("--type", required=True, choices=tuple(PARSERS))
    parser.add_argument("--format", required=True, choices=("xlsx", "csv"))
    parser.add_argument("--profile", default="generic")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    if args.profile != "generic":
        print("ERROR only the verified generic profile is supported in the Phase 17 baseline")
        return 1
    try:
        export(Path(args.input), args.type, args.format, Path(args.output), args.profile)
    except Exception as exc:
        print(f"ERROR {exc}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
